#!/usr/bin/env python3
"""
parsear_ajustes.py — Convierte la TABLA EXCEL DE AJUSTES del revisor
(nuevo flujo: ya no ingresa el Word del docente) al mismo JSON de
"distribución" que consume el tablero del migrador.

Estructura esperada del Excel (Hoja1), tolerante a filas/columnas extra:

    Código banner | ADMI_4038
    URL           | https://utpl.instructure.com/courses/11078

                  | Semana 1 | Semana 2 | ...
    RA            | 1        | 1, 2     | ...
    Unidades      | 1, 2     | 2        | ...
    Temas reubicados | 1.1., 1.2., ... | 2.3., 2.4. | ...
    Temas eliminados | 4.4. (Semana 13) | ...

    Consideraciones de eliminación:
        Re-enumeración de figuras y tablas, temas

Como el Excel solo trae NÚMEROS y CÓDIGOS, los títulos de temas/unidades
y los textos de los RA se completan desde el curso Canvas ya extraído
(enriquecer_con_canvas). Así el JSON final es idéntico en forma al que
producía parsear_distribucion.py y el front no necesita cambios de lógica:
todo tema del Canvas que no esté en el Excel queda automáticamente en
"Eliminar" (rojo) en el tablero.

Uso CLI (sin enriquecer):
    python3 parsear_ajustes.py Ajustes_ADMI_4038.xlsx [-o ajustes.json]
"""
import argparse, json, re
from openpyxl import load_workbook

RE_CODIGO = re.compile(r'\d+\.\d+(?:\.\d+)?')          # 1.1 / 2.5.3
RE_SEMANA_COL = re.compile(r'^semana\s+(\d+)\s*$', re.I)
RE_SEMANA_ORIGEN = re.compile(r'\(\s*semana\s+(\d+)\s*\)', re.I)
RE_NUMS = re.compile(r'\d+')


def _txt(v):
    return str(v).strip() if v is not None else ''


def _norm_codigo(c):
    """'1.1.' → '1.1' (los códigos del Excel traen punto final)."""
    return c.rstrip('.')


def parse(path):
    """Lee el Excel de ajustes → dict crudo (sin títulos ni textos de RA)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]  # Hoja1: la matriz de ajustes
    rows = [[_txt(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    meta = {'codigo_banner': '', 'curso_url': ''}
    col_semana = {}          # índice de columna → "Semana N"
    semanas = {}             # "Semana N" → {ra, unidades, temas}
    eliminados = []
    consideraciones = []
    en_consideraciones = False

    for row in rows:
        primera = next((c for c in row if c), '')
        etiqueta = row[0].lower() if row else ''

        # --- metadatos ---
        if 'banner' in etiqueta or 'banner' in primera.lower():
            vals = [c for c in row[1:] if c]
            if vals: meta['codigo_banner'] = vals[0]
            continue
        if etiqueta == 'url' or primera.lower() == 'url':
            vals = [c for c in row if c.startswith('http')]
            if vals: meta['curso_url'] = vals[0]
            continue

        # --- consideraciones de eliminación ---
        if 'consideraciones' in primera.lower():
            en_consideraciones = True
            continue
        if en_consideraciones:
            nota = ' '.join(c for c in row if c).strip()
            if nota: consideraciones.append(nota)
            continue

        # --- fila de encabezado con las semanas ---
        if any(RE_SEMANA_COL.match(c) for c in row):
            col_semana = {}
            for i, c in enumerate(row):
                m = RE_SEMANA_COL.match(c)
                if m:
                    nombre = f"Semana {int(m.group(1))}"
                    col_semana[i] = nombre
                    semanas.setdefault(nombre, {'ra': [], 'unidades': [], 'temas': []})
            continue

        if not col_semana:
            continue  # todavía no llegó la matriz

        et = etiqueta or primera.lower()

        if et.startswith('ra'):
            for i, sem in col_semana.items():
                if i < len(row) and row[i]:
                    semanas[sem]['ra'] = [int(n) for n in RE_NUMS.findall(row[i])]
        elif et.startswith('unidad'):
            for i, sem in col_semana.items():
                if i < len(row) and row[i]:
                    semanas[sem]['unidades'] = [int(n) for n in RE_NUMS.findall(row[i])]
        elif 'reubicad' in et or et.startswith('temas'):
            if 'eliminad' in et:
                # "Temas eliminados": pueden venir en cualquier columna
                for c in row[1:]:
                    for m in RE_CODIGO.finditer(c):
                        cod = _norm_codigo(m.group(0))
                        origen = RE_SEMANA_ORIGEN.search(c)
                        eliminados.append({
                            'codigo': cod,
                            'semana_origen': f"Semana {origen.group(1)}" if origen else None,
                        })
            else:
                for i, sem in col_semana.items():
                    if i < len(row) and row[i]:
                        semanas[sem]['temas'] = [
                            _norm_codigo(m.group(0)) for m in RE_CODIGO.finditer(row[i])
                        ]
        elif 'eliminad' in et:
            for c in row[1:]:
                for m in RE_CODIGO.finditer(c):
                    cod = _norm_codigo(m.group(0))
                    origen = RE_SEMANA_ORIGEN.search(c)
                    eliminados.append({
                        'codigo': cod,
                        'semana_origen': f"Semana {origen.group(1)}" if origen else None,
                    })

    return {'meta': meta, 'semanas': semanas,
            'eliminados': eliminados, 'consideraciones': consideraciones}


# ---------------------------------------------------------------------------
# Enriquecimiento desde el curso Canvas extraído
# ---------------------------------------------------------------------------
RE_HTML_RA = re.compile(
    r'Resultado de aprendizaje\s+(\d+)\s*:?\s*</strong>\s*</h2>\s*(?:</div>\s*)*<p>(.*?)</p>',
    re.S | re.I)
RE_HTML_UNIDAD = re.compile(
    r'<h3[^>]*>\s*(?:<strong>)?\s*Unidad\s+(\d+)\.\s*(.*?)(?:</strong>)?\s*</h3>', re.S | re.I)
RE_HTML_HEAD = re.compile(r'<(h[2-6])[^>]*>(.*?)</\1>', re.S | re.I)
RE_TAGS = re.compile(r'<[^>]+>')


def _indices_canvas(canvas):
    """Del curso extraído: {n_ra: texto}, {n_unidad: título}, {codigo: título}."""
    ras, unidades, temas = {}, {}, {}
    for m in canvas.get('modulos', []):
        for it in m.get('items', []):
            html = it.get('html') or ''
            for n, texto in RE_HTML_RA.findall(html):
                ras.setdefault(int(n), RE_TAGS.sub('', texto).replace('\xa0', ' ').strip())
            for n, titulo in RE_HTML_UNIDAD.findall(html):
                unidades.setdefault(int(n), RE_TAGS.sub('', titulo).replace('\xa0', ' ').strip())
            for _, contenido in RE_HTML_HEAD.findall(html):
                txt = RE_TAGS.sub('', contenido).replace('\xa0', ' ').strip()
                mm = re.match(r'^(\d+\.\d+(?:\.\d+)?)\.?\s+(.*)', txt)
                if mm:
                    temas.setdefault(mm.group(1), mm.group(2).strip())
    return ras, unidades, temas


def enriquecer_con_canvas(ajustes, canvas):
    """Ajustes crudos + curso Canvas → JSON de distribución (mismo formato
    que parsear_distribucion.py) para que el tablero funcione sin cambios."""
    ras_txt, uni_txt, temas_txt = _indices_canvas(canvas)
    dist = {'semanas': {}}
    for sem, d in ajustes['semanas'].items():
        unidades = []
        for u in d['unidades']:
            codigos = [c for c in d['temas'] if int(c.split('.')[0]) == u]
            unidades.append({
                'unidad': u,
                'titulo': uni_txt.get(u, ''),
                'marcada': False,   # el Excel no trae marca de "modificar"
                'temas': [{'codigo': c,
                           'titulo': temas_txt.get(c, ''),
                           'marcado': False} for c in codigos],
            })
        # temas cuyo dígito de unidad no esté en la fila Unidades (por si acaso)
        sueltos = [c for c in d['temas']
                   if int(c.split('.')[0]) not in d['unidades']]
        for c in sueltos:
            u = int(c.split('.')[0])
            destino = next((x for x in unidades if x['unidad'] == u), None)
            if not destino:
                destino = {'unidad': u, 'titulo': uni_txt.get(u, ''),
                           'marcada': False, 'temas': []}
                unidades.append(destino)
            destino['temas'].append({'codigo': c,
                                     'titulo': temas_txt.get(c, ''),
                                     'marcado': False})
        dist['semanas'][sem] = {
            'resultados_aprendizaje': [ras_txt.get(n, f'Resultado de aprendizaje {n}')
                                       for n in d['ra']],
            'unidades': unidades,
            'ra_unificados': len(d['ra']) > 1,
        }
    dist['eliminados'] = ajustes['eliminados']
    dist['consideraciones'] = ajustes['consideraciones']
    dist['codigo_banner'] = ajustes['meta']['codigo_banner']
    dist['curso_url'] = ajustes['meta']['curso_url']
    return dist


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('-o', '--out', default='ajustes.json')
    a = ap.parse_args()
    data = parse(a.xlsx)
    json.dump(data, open(a.out, 'w'), ensure_ascii=False, indent=2)
    total = sum(len(s['temas']) for s in data['semanas'].values())
    print(f"✓ {a.out} — {len(data['semanas'])} semanas, {total} temas reubicados, "
          f"{len(data['eliminados'])} eliminados")
